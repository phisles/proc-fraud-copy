import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re
from datetime import datetime

PROCESSED_DIRECTORY = "./processed_data"
CSV_DIRECTORY = "."  # Where the CSVs are stored

# Generate timestamp for the output Excel file
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_EXCEL_FILE = f"./formatted_report_{timestamp}.xlsx"

def find_latest_file(prefix):
    """Find the most recent file matching the given prefix."""
    csv_files = [f for f in os.listdir(CSV_DIRECTORY) if re.match(fr'{prefix}_\d{{8}}_\d{{4}}\.csv$', f)]
    if not csv_files:
        raise FileNotFoundError(f"No {prefix}_YYYYMMDD_HHMM.csv files found.")
    
    latest_file = max(csv_files, key=lambda x: x.split("_")[-1].split(".")[0])  # Sort by timestamp
    print(f"Using latest {prefix} file: {latest_file}")
    return os.path.join(CSV_DIRECTORY, latest_file)

def count_files_in_directory(directory):
    """Count the number of JSON files in the processed directory (representing original PDFs)."""
    return len([f for f in os.listdir(directory) if f.endswith(".json")])

def process_and_format_excel():
    """Create an Excel report with Summary, PDF Comparison, and Image Matches reports."""
    
    # Locate the latest CSV files
    pdf_comparison_csv = find_latest_file("pdf_comparison")
    image_matches_csv = find_latest_file("image_matches")
    
    # Load CSV data
    pdf_comparison_df = pd.read_csv(pdf_comparison_csv)
    image_matches_df = pd.read_csv(image_matches_csv)

    # Replace ".json" with ".pdf"
    pdf_comparison_df.replace(".json", ".pdf", regex=True, inplace=True)
    image_matches_df.replace(".json", ".pdf", regex=True, inplace=True)

    # Count total original files and matching files
    total_files = count_files_in_directory(PROCESSED_DIRECTORY)
    num_matches = len(pdf_comparison_df)
    num_matching_files = num_matches * 2  # Each row represents a pair, so 2 files per match
    match_percentage = (num_matching_files / total_files) * 100 if total_files > 0 else 0

    # Create Summary DataFrame
    summary_data = {
        "Metric": ["Total Original Files", "Number of Matches", "Matching File Percentage"],
        "Value": [total_files, num_matches, f"{match_percentage:.2f}%"]
    }
    summary_df = pd.DataFrame(summary_data)

    # Create Excel writer
    with pd.ExcelWriter(OUTPUT_EXCEL_FILE, engine="openpyxl") as writer:
        # Write Summary Report
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        workbook = writer.book
        summary_sheet = writer.sheets["Summary"]

        # Format Summary Sheet
        summary_sheet["A1"].font = Font(bold=True)
        summary_sheet["B1"].font = Font(bold=True)
        for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
        
        # Write PDF Comparison Report
        pdf_comparison_df.to_excel(writer, sheet_name="PDF Comparison", index=False)
        pdf_sheet = writer.sheets["PDF Comparison"]

        # Formatting PDF Comparison Sheet
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        for col in range(1, pdf_sheet.max_column + 1):
            pdf_sheet.cell(row=1, column=col).fill = header_fill
            pdf_sheet.cell(row=1, column=col).font = header_font
            pdf_sheet.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        # Write Image Matches Report
        image_matches_df.to_excel(writer, sheet_name="Image Matches", index=False)
        img_sheet = writer.sheets["Image Matches"]

        # Formatting Image Matches Sheet
        group_colors = ["FFD700", "87CEEB"]  # Alternating colors for groups
        color_index = 0
        current_group = None

        for row in range(2, img_sheet.max_row + 1):
            group_identifier = (img_sheet.cell(row=row, column=1).value, img_sheet.cell(row=row, column=2).value)

            if group_identifier != current_group:
                current_group = group_identifier
                color_index = (color_index + 1) % len(group_colors)

            fill_color = group_colors[color_index]
            for col in range(1, img_sheet.max_column + 1):
                img_sheet.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    print(f"Formatted Excel report saved to {OUTPUT_EXCEL_FILE}")

if __name__ == "__main__":
    process_and_format_excel()
