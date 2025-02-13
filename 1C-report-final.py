import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# File paths
summary_csv = "/Users/philip/Desktop/Code/proc-fraud-pdf/summary_report_with_contacts.csv"
matches_csv = "/Users/philip/Desktop/Code/proc-fraud-pdf/matches_report.csv"
output_excel = "/Users/philip/Desktop/Code/proc-fraud-pdf/formatted_report.xlsx"

# Load CSVs
summary_df = pd.read_csv(summary_csv)
matches_df = pd.read_csv(matches_csv)

# Create an Excel writer
with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    # Write Summary Report
    summary_df.to_excel(writer, sheet_name="Summary Report", index=False)
    workbook = writer.book
    summary_sheet = writer.sheets["Summary Report"]

    # Formatting Summary Report
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True)
    row_colors = ["FFFFFF", "F2F2F2"]  # Alternating row colors

    for col in range(1, summary_sheet.max_column + 1):
        summary_sheet.cell(row=1, column=col).fill = header_fill
        summary_sheet.cell(row=1, column=col).font = header_font
        summary_sheet.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    for row in range(2, summary_sheet.max_row + 1):
        fill_color = row_colors[row % 2]
        for col in range(1, summary_sheet.max_column + 1):
            summary_sheet.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Write Matches Report
    matches_df.to_excel(writer, sheet_name="Matches Report", index=False)
    matches_sheet = writer.sheets["Matches Report"]

    # Formatting Matches Report
    current_group = None
    group_colors = ["FFD700", "87CEEB"]  # Alternating colors for groups
    color_index = 0

    for row in range(2, matches_sheet.max_row + 1):
        group_identifier = (matches_sheet.cell(row=row, column=1).value, matches_sheet.cell(row=row, column=2).value)

        if group_identifier != current_group:
            current_group = group_identifier
            color_index = (color_index + 1) % len(group_colors)

        fill_color = group_colors[color_index]
        for col in range(1, matches_sheet.max_column + 1):
            matches_sheet.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Save the Excel file
    writer.book.save(output_excel)

print(f"Formatted Excel report saved to {output_excel}")